import tkinter as tk
from tkinter import *
from tkinter import filedialog
from pathlib import Path
import docx2txt

def click():
    Text_input_window.destroy()

Text_input_window= Tk()
label_1=Label (Text_input_window,text="Enter search word:", bg="black", fg="white").grid(row=1, column=0, sticky=W)

search_string= StringVar()
entry_1= Entry(Text_input_window, textvariable=search_string, width=20, bg="white").grid(row=2, column=0, sticky=W)
Button(Text_input_window, text="SUBMIT", width=6, command=click).grid(row=3, column=0, sticky=W)
Text_input_window.mainloop()

search_word= search_string.get()
search_list=[]
file_list=[]

root = Tk()
root.fileName = filedialog.askdirectory()
path = Path(root.fileName)
root.destroy()
root.mainloop()


for x in path.glob('*.docx'):
    try:
        paragraph_list = docx2txt.process(x).splitlines()
    except:
        paragraph_list = (f"searched for: {search_word} but file could not be searched due to formatting")

    for i in paragraph_list:
        if i.find(search_word)>=0:
            search_list.append(i)
            file_list.append(x)

print(search_list)
print(file_list)

InfoWindow = Tk()

InfoWindow.title("Searched for:")
Close_Button = Button(InfoWindow,text="Close",command=InfoWindow.destroy)
Close_Button.grid(row=len(search_list)+3, column=2, sticky='w')
Cl_Heading_1 = tk.Label(InfoWindow,text="Files:", font=("arial",10, "underline", "bold")).grid(row=1, column=0, sticky='w')
Cl_Heading_2 = tk.Label(InfoWindow,text="Paragraph word found in:", font=("arial",10, "underline", "bold")).grid(row=1, column=4, sticky='w')

file_list_dict = {}
search_list_dict={}
for i in range(len(file_list)):
    file_list_dict[i] = tk.Label(InfoWindow,text=file_list[i]).grid(row=2+i, column=0, sticky='w')
    for y in range(len(search_list)):
        search_list_dict[y] = tk.Label(InfoWindow, text=search_list[y]).grid(row=2+y, column=4, sticky='w')


InfoWindow.mainloop()

counter=1
a=0
import openpyxl as xl
wb = xl.load_workbook(r"D:\Documents\Nick test folder\Search output.xlsx")
sheet = wb['Sheet1']
while a <= len(search_list)-1:
    #print(search_list[a])
    sheet.cell(counter,1).value=search_list[a]
    sheet.cell(counter, 2).value = str(file_list[a])
    counter+=1
    a +=1
wb.save(r"D:\Documents\Nick test folder\Search output.xlsx")
